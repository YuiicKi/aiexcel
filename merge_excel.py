import pandas as pd
import os
import time
from datetime import datetime
from openai import OpenAI
import openpyxl
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from compare_headers import read_excel_headers, compare_headers_with_ai
from mapping_cache import MappingCache

# 特殊列配置
SPECIAL_COLUMNS = {
    'student_id': 3,  # 学号列
    'signature': -1,  # 签名列（最后一列）
}

def format_date_with_ai(dates):
    # 初始化API客户端
    client = OpenAI(
        api_key="***",
        base_url="***"
    )

    # 构建提示信息
    dates_str = "\n".join([f"{i+1}. {d}" for i, d in enumerate(dates) if pd.notna(d)])
    prompt = f"""请将以下日期统一转换为mm.dd格式（月份和日期都用两位数表示，中间用点分隔）。
如果日期明显错误或无法解析，请返回空字符串。
示例：
1月7日 -> 01.07
1/7 -> 01.07
2024/1/7 -> 01.07
2024-01-07 -> 01.07

需要转换的日期：
{dates_str}

请按照以下格式返回结果，每行一个日期：
01.07
02.15
...
只返回转换后的日期，不要有任何解释性文字。对于空值或无法解析的日期，返回空行。"""

    try:
        print("正在统一日期格式...")
        response = client.chat.completions.create(
            model="***",
            messages=[
                {"role": "system", "content": "你是一个专门处理日期格式的助手。请直接返回格式化后的日期，不要返回其他内容。"},
                {"role": "user", "content": prompt}
            ],
            stream=True
        )

        # 收集完整的响应
        full_response = ""
        for chunk in response:
            if hasattr(chunk.choices[0].delta, 'content') and chunk.choices[0].delta.content is not None:
                content = chunk.choices[0].delta.content
                full_response += content

        # 解析返回的日期
        formatted_dates = full_response.strip().split('\n')
        
        # 创建结果列表，保持原始数据的空值
        result = []
        formatted_idx = 0
        for original_date in dates:
            if pd.isna(original_date):
                result.append('')
            else:
                if formatted_idx < len(formatted_dates):
                    result.append(formatted_dates[formatted_idx].strip())
                    formatted_idx += 1
                else:
                    result.append('')
        
        return result
    except Exception as e:
        print(f"日期格式化错误：{str(e)}")
        return dates

def parse_ai_response(response_text):
    """解析AI响应文本，转换为字典映射"""
    mapping = {}
    for line in response_text.strip().split('\n'):
        if '对应' in line:
            parts = line.replace('。', '').split('对应')
            table1_part = parts[0].strip()
            table2_part = parts[1].strip()
            
            # 提取索引
            idx1 = int(table1_part.split('的')[1]) - 1
            idx2 = int(table2_part.split('的')[1]) - 1
            
            # 创建映射关系（从table2到table1）
            mapping[idx2] = idx1
    return mapping

def insert_signature_images(ws, signature_col, image_dir):
    """
    将签名列中的文本替换为对应的图片
    
    参数：
        ws：openpyxl工作表对象
        signature_col：签名列索引（1-based，负数表示从末尾开始计数）
        image_dir：图片目录路径
    """
    from openpyxl.drawing.image import Image
    import os
    
    # 获取数据范围
    data_rows = ws.max_row
    data_cols = ws.max_column
    
    # 处理负数列索引
    if signature_col < 0:
        signature_col = data_cols + signature_col + 1
    
    # 获取图片目录中的所有文件
    image_files = os.listdir(image_dir)
    
    # 获取默认单元格大小（以像素为单位）
    default_row_height = 20  # Excel默认行高（约20像素）
    default_col_width = 64   # Excel默认列宽（约64像素）
    
    # 遍历签名列中的每个单元格
    for row in range(2, data_rows + 1):  # 从第二行开始（跳过标题行）
        cell = ws.cell(row=row, column=signature_col)
        if cell.value:
            # 构造完整的图片文件名模式（包括时间戳）
            cell_value = str(cell.value).strip()
            
            # 在图片目录中查找匹配的图片文件
            matching_image = None
            for image_file in image_files:
                # 检查文件名是否包含单元格值（考虑时间戳等）
                if cell_value in image_file and any(image_file.lower().endswith(ext) for ext in ['.png', '.jpg', '.jpeg']):
                    matching_image = image_file
                    break
            
            if matching_image:
                try:
                    # 创建图片对象
                    image_path = os.path.join(image_dir, matching_image)
                    img = Image(image_path)
                    
                    # 计算缩放比例以适应单元格大小
                    width_ratio = default_col_width / img.width
                    height_ratio = default_row_height / img.height
                    scale_ratio = min(width_ratio, height_ratio)
                    
                    # 缩放图片以适应单元格
                    img.width = int(img.width * scale_ratio)
                    img.height = int(img.height * scale_ratio)
                    
                    # 将图片添加到单元格
                    img.anchor = f"{get_column_letter(signature_col)}{row}"
                    ws.add_image(img)
                    
                    # 清除单元格文本内容
                    cell.value = None
                    
                except Exception as e:
                    print(f"处理图片{matching_image}出错：{str(e)}")
            else:
                print(f"没有找到匹配的图片：{cell_value}")

def apply_excel_formatting(output_path):
    # 打开工作簿
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    
    # 定义边框样式（正常边框）
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 获取数据范围
    data_rows = ws.max_row
    data_cols = ws.max_column
    
    # 合并第一行单元格（从A1到最后一列）
    ws.merge_cells(f'A1:{get_column_letter(data_cols)}1')
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    
    # 遍历所有单元格
    for row in range(1, data_rows + 1):
        for col in range(1, data_cols + 1):
            cell = ws.cell(row=row, column=col)
            
            # 添加边框和居中对齐到所有单元格
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 如果是学号列，设置为文本格式
            if row > 1 and col == SPECIAL_COLUMNS['student_id']:  
                cell.number_format = '@'
    
    # 设置第一行高度
    ws.row_dimensions[1].height = 30  # 设置为30单位
    
    # 处理签名图片（假设签名在最后一列）
    image_dir = "***"
    if os.path.exists(image_dir):
        insert_signature_images(ws, SPECIAL_COLUMNS['signature'], image_dir)
    
    # 保存更改
    wb.save(output_path)

def merge_excel_files():
    start_time = time.time()
    print("正在合并Excel文件...")
    # 初始化缓存系统
    cache = MappingCache()
    
    # 读取三个Excel文件
    file1_path = "***"
    file2_path = "***"
    file3_path = "***"
    
    # 读取表1（包括第一行）、表2和表3
    df1_full = pd.read_excel(file1_path)  # 读取完整的表1，包括第一行
    df1 = pd.read_excel(file1_path, header=1)  # 读取表1的数据部分
    df2 = pd.read_excel(file2_path, dtype={'请输入你的学号（必填）': str})  # 读取表2，学号列作为字符串
    df3 = pd.read_excel(file3_path, header=1)  # 读取表3，跳过第一行
    
    # 获取标题
    headers1 = df1.columns.tolist()
    headers2 = df2.columns.tolist()
    headers3 = df3.columns.tolist()
    
    # 尝试从缓存中获取表1和表2之间的映射关系
    index_mapping = cache.get_mapping(headers1, headers2, "1_to_2")
    if index_mapping is None:
        print("没有找到缓存中的映射关系，使用AI分析...")
        # 使用AI获取表1和表2之间的列映射关系
        ai_response = compare_headers_with_ai(headers1, headers2)
        index_mapping = parse_ai_response(ai_response)
        if index_mapping:
            # 保存映射关系到缓存
            cache.save_mapping(headers1, headers2, "1_to_2", index_mapping)
    else:
        print("使用缓存中的映射关系")
    
    if not index_mapping:
        print("获取列映射关系失败，程序终止")
        return

    # 尝试从缓存中获取表1和表3之间的映射关系
    table3_mapping = cache.get_mapping(headers1, headers3, "1_to_3")
    if table3_mapping is None:
        print("没有找到缓存中的映射关系，使用AI分析...")
        # 使用AI获取表1和表3之间的列映射关系
        ai_response3 = compare_headers_with_ai(headers1, headers3, is_comparing_2_and_3=False, is_comparing_1_and_3=True)
        if ai_response3:
            table3_mapping = {}
            for line in ai_response3.strip().split('\n'):
                if '对应' in line:
                    parts = line.replace('。', '').split('对应')
                    table1_part = parts[0].strip()
                    table3_part = parts[1].strip()
                    
                    # 提取索引（转换为0-based索引）
                    idx1 = int(table1_part.split('的')[1]) - 1
                    idx3 = int(table3_part.split('的')[1]) - 1
                    
                    # 创建映射关系（从表3到表1）
                    table3_mapping[idx3] = idx1
            
            if table3_mapping:
                # 保存映射关系到缓存
                cache.save_mapping(headers1, headers3, "1_to_3", table3_mapping)
    else:
        print("使用缓存中的映射关系")
    
    if not table3_mapping:
        print("获取列映射关系失败，程序终止")
        return

    # 创建一个新的DataFrame，完全复制表1的结构
    new_df = pd.DataFrame(columns=headers1)
    
    # 首先填充序号列（如果存在）
    if '序号' in headers1:
        new_df['序号'] = range(1, len(df2) + 1)
    
    # 根据映射关系填充表2的数据
    for idx2, idx1 in index_mapping.items():
        # 获取表2的数据并填充到表1对应的列中
        new_df[headers1[idx1]] = df2[headers2[idx2]]
    
    # 处理学号列
    if '学号' in new_df.columns:
        new_df['学号'] = new_df['学号'].astype(str)
    
    # 处理日期格式
    if '离校时间' in new_df.columns:
        new_df['离校时间'] = format_date_with_ai(new_df['离校时间'].tolist())
    if '返校时间' in new_df.columns:
        new_df['返校时间'] = format_date_with_ai(new_df['返校时间'].tolist())
    
    # 合并表3的数据
    print("\n合并表3的数据...")
    print("表1和表3的对应关系：")
    for idx3, idx1 in table3_mapping.items():
        print(f"表3的第{idx3+1}列 -> 表1的第{idx1+1}列")
    
    # 确保学号列存在于new_df中
    if '学号' in new_df.columns:
        # 获取表3的学号列（根据映射关系）
        student_id_col = None
        for idx3, idx1 in table3_mapping.items():
            if headers1[idx1] == '学号':
                student_id_col = idx3
                break
        
        if student_id_col is not None:
            df3['学号'] = df3.iloc[:, student_id_col].astype(str)
            
            # 遍历表3的每一行
            for idx3, row3 in df3.iterrows():
                # 在new_df中查找匹配的学号
                mask = new_df['学号'] == row3['学号']
                if mask.any():
                    # 找到匹配的行，更新对应的列
                    for col3, col1 in table3_mapping.items():
                        new_df.loc[mask, headers1[col1]] = row3.iloc[col3]
    
    # 按学号排序
    if '学号' in new_df.columns:
        new_df = new_df.sort_values('学号')
        # 重新生成序号
        if '序号' in new_df.columns:
            new_df['序号'] = range(1, len(new_df) + 1)
    
    # 生成新的文件名（包括时间戳）
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_filename = f'Merged_result_{timestamp}.xlsx'
    output_path = os.path.join(os.path.dirname(file1_path), output_filename)
    
    # 获取第一行数据
    first_row = pd.read_excel(file1_path, nrows=1)
    
    # 使用ExcelWriter保存，允许多个DataFrame
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 首先写入第一行
        first_row.to_excel(writer, index=False)
        # 写入数据，从第二行开始
        new_df.to_excel(writer, index=False, startrow=1)
    
    # 应用格式化
    apply_excel_formatting(output_path)
    
    print(f"\n生成的合并文件：{output_filename}")
    print(f"包含{len(new_df)}条记录")
    print("应用了格式化：")
    print("- 保留了原始表的第一行和合并单元格")
    print("- 居中对齐了所有单元格")
    print("- 合并了表2和表3的数据")
    print("- 添加了细边框到所有单元格")
    print("- 将学号列设置为文本格式以避免科学计数法")
    print("- 统一了日期格式为mm.dd")
    print(f"\nExcel合并完成！总耗时：{time.time() - start_time:.2f}秒")
    return output_path

if __name__ == "__main__":
    merge_excel_files()
